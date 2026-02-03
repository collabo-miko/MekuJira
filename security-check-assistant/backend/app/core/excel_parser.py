"""Excel file parser for extracting questions."""
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from dataclasses import dataclass

from app.models.question import Question, QuestionType


@dataclass
class SheetFormat:
    """Detected format of an Excel sheet."""

    sheet_name: str
    question_column: str
    answer_column: str
    remarks_column: Optional[str]
    header_row: int
    data_start_row: int
    confidence: float


@dataclass
class ExtractedQuestion:
    """Question extracted from Excel."""

    row_number: int
    question_text: str
    remarks: Optional[str]
    answer_column: str
    question_type: QuestionType
    choices: Optional[list[str]]


class ExcelParser:
    """Parser for security check sheet Excel files."""

    def __init__(self, file_path: str):
        """Initialize parser with file path."""
        self.file_path = Path(file_path)
        self._workbook: Optional[Workbook] = None

    def load(self) -> None:
        """Load the Excel file."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")
        self._workbook = load_workbook(str(self.file_path), data_only=True)

    @property
    def workbook(self) -> Workbook:
        """Get the loaded workbook."""
        if self._workbook is None:
            self.load()
        return self._workbook

    def get_sheet_names(self) -> list[str]:
        """Get all sheet names."""
        return self.workbook.sheetnames

    def extract_sample_data(
        self,
        sheet_name: Optional[str] = None,
        max_rows: int = 15,
        max_cols: int = 10,
    ) -> list[list[str]]:
        """Extract sample data from a sheet for format detection."""
        ws = self.workbook[sheet_name] if sheet_name else self.workbook.active

        data = []
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols), 1):
            row_data = []
            for cell in row:
                value = cell.value
                if value is not None:
                    row_data.append(str(value)[:100])  # Truncate long values
                else:
                    row_data.append("")
            data.append(row_data)

        return data

    def extract_questions(
        self,
        sheet_format: SheetFormat,
        sheet_name: Optional[str] = None,
    ) -> list[ExtractedQuestion]:
        """Extract questions based on detected format."""
        ws = self.workbook[sheet_name] if sheet_name else self.workbook.active

        questions = []
        q_col_idx = self._column_to_index(sheet_format.question_column)
        a_col_idx = self._column_to_index(sheet_format.answer_column)
        r_col_idx = (
            self._column_to_index(sheet_format.remarks_column)
            if sheet_format.remarks_column
            else None
        )

        for row_idx, row in enumerate(ws.iter_rows(min_row=sheet_format.data_start_row), sheet_format.data_start_row):
            # Get question text
            question_cell = row[q_col_idx] if q_col_idx < len(row) else None
            if question_cell is None or question_cell.value is None:
                continue

            question_text = str(question_cell.value).strip()
            if not question_text:
                continue

            # Get remarks if available
            remarks = None
            if r_col_idx is not None and r_col_idx < len(row):
                remarks_cell = row[r_col_idx]
                if remarks_cell and remarks_cell.value:
                    remarks = str(remarks_cell.value).strip()

            # Detect question type
            question_type, choices = self._detect_question_type(question_text, row, a_col_idx)

            questions.append(
                ExtractedQuestion(
                    row_number=row_idx,
                    question_text=question_text,
                    remarks=remarks,
                    answer_column=sheet_format.answer_column,
                    question_type=question_type,
                    choices=choices,
                )
            )

        return questions

    def _column_to_index(self, column: str) -> int:
        """Convert column letter to 0-based index."""
        result = 0
        for char in column.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result - 1

    def _detect_question_type(
        self,
        question_text: str,
        row,
        answer_col_idx: int,
    ) -> tuple[QuestionType, Optional[list[str]]]:
        """Detect the type of question and extract choices if applicable."""
        question_lower = question_text.lower()

        # Check for yes/no type keywords
        yes_no_keywords = ["はい", "いいえ", "yes", "no", "該当", "非該当"]
        if any(kw in question_lower for kw in yes_no_keywords):
            return QuestionType.YES_NO, ["はい", "いいえ"]

        # Check for multiple choice indicators
        if "以下から選択" in question_text or "選んでください" in question_text:
            return QuestionType.MULTIPLE_CHOICE, None

        # Default to free text
        return QuestionType.FREE_TEXT, None

    def write_answers(
        self,
        sheet_format: SheetFormat,
        answers: dict[int, str],
        output_path: Optional[str] = None,
        sheet_name: Optional[str] = None,
    ) -> str:
        """Write answers back to the Excel file.

        Args:
            sheet_format: The detected format
            answers: Dict mapping row_number to answer text
            output_path: Optional output path (defaults to original with _filled suffix)
            sheet_name: Optional sheet name

        Returns:
            Path to the output file
        """
        ws = self.workbook[sheet_name] if sheet_name else self.workbook.active
        a_col_idx = self._column_to_index(sheet_format.answer_column) + 1  # 1-based for openpyxl

        for row_number, answer_text in answers.items():
            ws.cell(row=row_number, column=a_col_idx, value=answer_text)

        # Determine output path
        if output_path is None:
            output_path = str(self.file_path.with_stem(f"{self.file_path.stem}_filled"))

        self.workbook.save(output_path)
        return output_path

    def close(self) -> None:
        """Close the workbook."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
