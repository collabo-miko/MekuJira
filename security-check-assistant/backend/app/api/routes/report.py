"""Report and export API routes."""
import io
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import load_workbook

from app.config import get_settings
from app.core.excel_parser import ExcelParser, SheetFormat
from app.db.repositories.session_repository import SessionRepository
from app.db.repositories.question_repository import QuestionRepository
from app.db.repositories.answer_repository import AnswerRepository
from app.db.repositories.knowledge_repository import KnowledgeRepository
from app.models.answer import AnswerStatus

router = APIRouter()


class ReportSummary(BaseModel):
    """Summary statistics for a session."""

    session_id: str
    filename: str
    vendor_name: Optional[str]
    total_questions: int
    answered_count: int
    high_confidence_count: int
    medium_confidence_count: int
    low_confidence_count: int
    approved_count: int
    modified_count: int
    pending_count: int


class DifficultQuestion(BaseModel):
    """A question that was difficult to answer."""

    question_id: str
    row_number: int
    question_text: str
    confidence_score: float
    reason: str


class ReportResponse(BaseModel):
    """Full report response."""

    summary: ReportSummary
    difficult_questions: list[DifficultQuestion]
    knowledge_count: int


@router.get("/report/{session_id}", response_model=ReportResponse)
async def get_report(session_id: str):
    """Get a detailed report for a session."""
    session = await SessionRepository.get_by_id(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")

    questions = await QuestionRepository.get_by_session(session_id)
    answers = await AnswerRepository.get_by_session(session_id)
    knowledge_count = await KnowledgeRepository.count()

    # Create answer lookup
    answer_map = {a.question_id: a for a in answers}

    # Calculate statistics
    answered_count = 0
    high_confidence_count = 0
    medium_confidence_count = 0
    low_confidence_count = 0
    approved_count = 0
    modified_count = 0
    pending_count = 0
    difficult_questions = []

    threshold = session.confidence_threshold

    for q in questions:
        answer = answer_map.get(q.id)
        if not answer:
            pending_count += 1
            difficult_questions.append(
                DifficultQuestion(
                    question_id=q.id,
                    row_number=q.row_number,
                    question_text=q.question_text,
                    confidence_score=0.0,
                    reason="回答が生成されていません",
                )
            )
            continue

        if answer.answer_text:
            answered_count += 1

        # Confidence levels
        if answer.confidence_score >= 0.85:
            high_confidence_count += 1
        elif answer.confidence_score >= 0.70:
            medium_confidence_count += 1
        else:
            low_confidence_count += 1

        # Status counts
        if answer.status == AnswerStatus.APPROVED:
            approved_count += 1
        elif answer.status == AnswerStatus.MODIFIED:
            modified_count += 1
        else:
            pending_count += 1

        # Difficult questions (below threshold or no answer)
        if answer.confidence_score < threshold or not answer.answer_text:
            reason = "確信度が閾値未満です" if answer.answer_text else "回答が見つかりませんでした"
            difficult_questions.append(
                DifficultQuestion(
                    question_id=q.id,
                    row_number=q.row_number,
                    question_text=q.question_text,
                    confidence_score=answer.confidence_score,
                    reason=reason,
                )
            )

    summary = ReportSummary(
        session_id=session_id,
        filename=session.filename,
        vendor_name=session.vendor_name,
        total_questions=len(questions),
        answered_count=answered_count,
        high_confidence_count=high_confidence_count,
        medium_confidence_count=medium_confidence_count,
        low_confidence_count=low_confidence_count,
        approved_count=approved_count,
        modified_count=modified_count,
        pending_count=pending_count,
    )

    return ReportResponse(
        summary=summary,
        difficult_questions=difficult_questions,
        knowledge_count=knowledge_count,
    )


@router.get("/export/{session_id}/excel")
async def export_excel(session_id: str):
    """Export the session with answers filled in to Excel."""
    session = await SessionRepository.get_by_id(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")

    settings = get_settings()
    file_path = Path(settings.uploads_dir) / session.filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Original Excel file not found")

    # Get questions and answers
    questions = await QuestionRepository.get_by_session(session_id)
    answers = await AnswerRepository.get_by_session(session_id)

    # Create answer lookup by question_id
    answer_map = {a.question_id: a for a in answers}

    # Create row_number -> answer_text mapping
    row_answers = {}
    for q in questions:
        answer = answer_map.get(q.id)
        if answer and answer.answer_text:
            row_answers[q.row_number] = answer.answer_text

    # Load workbook and write answers
    workbook = load_workbook(str(file_path))
    ws = workbook.active

    # Get answer column index
    if session.answer_column:
        # Convert column letter to number
        col_idx = 0
        for char in session.answer_column.upper():
            col_idx = col_idx * 26 + (ord(char) - ord("A") + 1)
    else:
        col_idx = 2  # Default to column B

    # Write answers
    for row_num, answer_text in row_answers.items():
        ws.cell(row=row_num, column=col_idx, value=answer_text)

    # Save to bytes buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Generate filename
    output_filename = f"{Path(session.filename).stem}_filled.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{output_filename}"'
        },
    )
